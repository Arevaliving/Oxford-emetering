#!/usr/bin/env python3
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE  = Path("/home/oxford/oxford-sync")
LJSON = BASE/"www/data/latest.json"
DJSON = BASE/"www/data/daily_history.json"
OUT   = BASE/"www/data/oxford_water_report.xlsx"

C_NAV="1A3A5C"; C_WHT="FFFFFF"; C_DHW="FF6B35"; C_DCW="4A9EBF"
C_ALT="F0F4F8"; C_TOT="D9E8F5"; C_PAR="FFF3CD"
C_DHP="E8A87C"; C_DCP="7ABDD4"
THIN=Side(style="thin",color="CCCCCC")
BRD=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
MONTHS=["Oct 25","Nov 25","Dec 25","Jan 26","Feb 26","Mar 26"]
PARTIAL_IDX=5

def mkfill(h): return PatternFill("solid",start_color=h,end_color=h)
def hf(): return Font(name="Arial",bold=True,color=C_WHT,size=10)
def bf(bold=False): return Font(name="Arial",bold=bold,size=10)
def ca(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def la(): return Alignment(horizontal="left",vertical="center")

def hc(ws,r,c,v,bg=None):
    x=ws.cell(r,c,v); x.font=hf(); x.fill=mkfill(bg or C_NAV)
    x.alignment=ca(); x.border=BRD; return x

def bc(ws,r,c,v,bold=False,fmt=None,bg=None,ctr=False):
    x=ws.cell(r,c,v); x.font=bf(bold)
    x.alignment=ca() if ctr else la(); x.border=BRD
    if fmt: x.number_format=fmt
    if bg: x.fill=mkfill(bg)
    return x

def ids(ws,units,sr):
    for i,u in enumerate(units):
        r=sr+i; bg=C_ALT if i%2 else None
        bc(ws,r,1,int(u["u"]),fmt="0",bg=bg,ctr=True)
        bc(ws,r,2,int(u["f"]),fmt="0",bg=bg,ctr=True)
        bc(ws,r,3,u.get("dh_s",""),bg=bg,ctr=True)
        bc(ws,r,4,u.get("dc_s",""),bg=bg,ctr=True)

def totrow(ws,cols,sr,n):
    last=sr+n-1; tr=last+1
    ws.merge_cells(f"A{tr}:D{tr}")
    t=ws[f"A{tr}"]; t.value="TOTAL"; t.font=bf(True)
    t.fill=mkfill(C_TOT); t.alignment=ca(); t.border=BRD
    for col in cols:
        cl=get_column_letter(col)
        c=ws.cell(tr,col,f"=SUM({cl}{sr}:{cl}{last})")
        c.font=bf(True); c.fill=mkfill(C_TOT)
        c.number_format="0.000"; c.alignment=ca(); c.border=BRD

def tab1(wb,units,gen):
    ws=wb.create_sheet("Current Readings"); ws.freeze_panes="E2"
    ws.merge_cells("A1:F1"); t=ws["A1"]
    t.value=f"Oxford Suites - Current Meter Readings  (as of {gen})"
    t.font=Font(name="Arial",bold=True,size=12,color=C_WHT)
    t.fill=mkfill(C_NAV); t.alignment=la()
    for col,h,bg in [(1,"Unit",C_NAV),(2,"Floor",C_NAV),(3,"DHW Serial",C_NAV),(4,"DCW Serial",C_NAV),(5,"DHW Cumulative (m3)",C_DHW),(6,"DCW Cumulative (m3)",C_DCW)]:
        hc(ws,2,col,h,bg=bg)
    ids(ws,units,3)
    for i,u in enumerate(units):
        r=3+i; bg=C_ALT if i%2 else None
        bc(ws,r,5,u.get("dh_cur") or None,fmt="#,##0.000",bg=bg,ctr=True)
        bc(ws,r,6,u.get("dc_cur") or None,fmt="#,##0.000",bg=bg,ctr=True)
    totrow(ws,[5,6],3,len(units))
    for col,w in [(1,8),(2,7),(3,16),(4,16),(5,22),(6,22)]:
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.row_dimensions[1].height=22; ws.row_dimensions[2].height=32

def tab2(wb,units):
    ws=wb.create_sheet("Daily Consumption"); ws.freeze_panes="E3"
    dates=sorted({d for u in units for d in u.get("daily",{}).keys()})
    nc=4+len(dates)*2
    ws.merge_cells(f"A1:{get_column_letter(nc)}1"); t=ws["A1"]
    t.value="Oxford Suites - Daily Consumption (m3/day)"
    t.font=Font(name="Arial",bold=True,size=12,color=C_WHT)
    t.fill=mkfill(C_NAV); t.alignment=la()
    for col,h in [(1,"Unit"),(2,"Floor"),(3,"DHW Serial"),(4,"DCW Serial")]:
        hc(ws,2,col,h)
    for j,d in enumerate(dates):
        lbl=datetime.strptime(d,"%Y-%m-%d").strftime("%b %d")
        hc(ws,2,5+j*2,f"DHW\n{lbl}",bg=C_DHW)
        hc(ws,2,6+j*2,f"DCW\n{lbl}",bg=C_DCW)
    ids(ws,units,3)
    for i,u in enumerate(units):
        r=3+i; bg=C_ALT if i%2 else None; dy=u.get("daily",{})
        for j,d in enumerate(dates):
            v=dy.get(d,{}); bc(ws,r,5+j*2,v.get("dh") or None,fmt="0.000",bg=bg,ctr=True)
            bc(ws,r,6+j*2,v.get("dc") or None,fmt="0.000",bg=bg,ctr=True)
    totrow(ws,[5+j*2 for j in range(len(dates))]+[6+j*2 for j in range(len(dates))],3,len(units))
    for col,w in [(1,7),(2,6),(3,15),(4,15)]: ws.column_dimensions[get_column_letter(col)].width=w
    for j in range(len(dates)):
        ws.column_dimensions[get_column_letter(5+j*2)].width=9
        ws.column_dimensions[get_column_letter(6+j*2)].width=9
    ws.row_dimensions[1].height=22; ws.row_dimensions[2].height=36

def tab3(wb,units,gen):
    ws=wb.create_sheet("Monthly Billing"); ws.freeze_panes="E3"
    nm=len(MONTHS); nc=4+nm*2
    ws.merge_cells(f"A1:{get_column_letter(nc)}1"); t=ws["A1"]
    t.value=f"Oxford Suites - Monthly Consumption by Billing Period  (updated {gen})"
    t.font=Font(name="Arial",bold=True,size=12,color=C_WHT)
    t.fill=mkfill(C_NAV); t.alignment=la()
    for col,h in [(1,"Unit"),(2,"Floor"),(3,"DHW Serial"),(4,"DCW Serial")]:
        hc(ws,2,col,h)
    for mi,lbl in enumerate(MONTHS):
        p=(mi==PARTIAL_IDX); disp=f"{lbl} \u2605" if p else lbl
        hc(ws,2,5+mi*2,f"DHW\n{disp}",bg=C_DHP if p else C_DHW)
        hc(ws,2,6+mi*2,f"DCW\n{disp}",bg=C_DCP if p else C_DCW)
    ids(ws,units,3)
    for i,u in enumerate(units):
        r=3+i; dha=u.get("dh") or []; dca=u.get("dc") or []
        for mi in range(nm):
            p=(mi==PARTIAL_IDX); rbg=C_PAR if p else (C_ALT if i%2 else None)
            bc(ws,r,5+mi*2,dha[mi] if mi<len(dha) and dha[mi] else None,fmt="0.000",bg=rbg,ctr=True)
            bc(ws,r,6+mi*2,dca[mi] if mi<len(dca) and dca[mi] else None,fmt="0.000",bg=rbg,ctr=True)
    totrow(ws,[5+mi*2 for mi in range(nm)]+[6+mi*2 for mi in range(nm)],3,len(units))
    for col,w in [(1,7),(2,6),(3,15),(4,15)]: ws.column_dimensions[get_column_letter(col)].width=w
    for mi in range(nm):
        ws.column_dimensions[get_column_letter(5+mi*2)].width=11
        ws.column_dimensions[get_column_letter(6+mi*2)].width=11
    ws.row_dimensions[1].height=22; ws.row_dimensions[2].height=36

def main():
    lat=json.loads(LJSON.read_text())
    units=sorted(lat["units"],key=lambda u:(int(u["f"]),int(u["u"])))
    gen=lat.get("last_sync","")[:16].replace("T"," ")+" UTC"
    wb=Workbook(); wb.remove(wb.active)
    tab1(wb,units,gen)
    tab2(wb,units)
    tab3(wb,units,gen)
    OUT.parent.mkdir(parents=True,exist_ok=True)
    wb.save(str(OUT))
    print(f"Saved {OUT}  ({OUT.stat().st_size:,} bytes)")

if __name__=="__main__":
    main()
